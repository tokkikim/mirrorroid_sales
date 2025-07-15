import os
import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns

from models.models import ReferenceArea, RecommendationResult
from config import settings

class ReportService:
    def __init__(self, db: Session):
        self.db = db
        self.reports_dir = settings.REPORT_FOLDER
        os.makedirs(self.reports_dir, exist_ok=True)
    
    async def generate_excel_report(
        self,
        reference_area: ReferenceArea,
        recommendations: List[RecommendationResult],
        include_charts: bool = True,
        include_details: bool = True
    ) -> str:
        """Excel 보고서 생성"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"franchise_recommendation_{reference_area.name}_{timestamp}.xlsx"
        file_path = os.path.join(self.reports_dir, file_name)
        
        # 데이터 준비
        recommendation_data = []
        for i, rec in enumerate(recommendations, 1):
            recommendation_data.append({
                '순위': i,
                '지역명': rec.recommended_location.name,
                '주소': rec.recommended_location.address,
                '유사도 점수': float(rec.similarity_score),
                '추천 사유': rec.recommendation_reason,
                '인구수': rec.recommended_location.population_total or 0,
                '임대료': rec.recommended_location.rent_price or 0,
                '경쟁업체 수': rec.recommended_location.competitor_count or 0,
                '교통 점수': rec.recommended_location.transportation_score or 0,
                '검토 상태': '검토 완료' if rec.is_reviewed else '검토 대기'
            })
        
        df = pd.DataFrame(recommendation_data)
        
        # Excel 파일 생성
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 요약 시트
            self._create_summary_sheet(writer, reference_area, recommendations)
            
            # 추천 결과 시트
            df.to_excel(writer, sheet_name='추천 결과', index=False)
            
            # 차트 시트 (옵션)
            if include_charts:
                self._create_charts_sheet(writer, df)
            
            # 상세 정보 시트 (옵션)
            if include_details:
                self._create_details_sheet(writer, reference_area, recommendations)
        
        # 스타일 적용
        self._apply_excel_styles(file_path)
        
        return file_path
    
    def _create_summary_sheet(self, writer, reference_area, recommendations):
        """요약 시트 생성"""
        summary_data = {
            '항목': [
                '기준 상권명',
                '기준 상권 주소',
                '월 매출',
                '상권 유형',
                '분석 일시',
                '총 추천 지역 수',
                '평균 유사도',
                '최고 유사도',
                '검토 완료율'
            ],
            '값': [
                reference_area.name,
                reference_area.address,
                f"{reference_area.monthly_sales:,}원" if reference_area.monthly_sales else 'N/A',
                reference_area.area_type or 'N/A',
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                len(recommendations),
                f"{sum(float(r.similarity_score) for r in recommendations) / len(recommendations):.3f}" if recommendations else 'N/A',
                f"{max(float(r.similarity_score) for r in recommendations):.3f}" if recommendations else 'N/A',
                f"{sum(1 for r in recommendations if r.is_reviewed) / len(recommendations) * 100:.1f}%" if recommendations else 'N/A'
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='요약', index=False)
    
    def _create_charts_sheet(self, writer, df):
        """차트 시트 생성"""
        # 간단한 차트 데이터 준비
        chart_data = pd.DataFrame({
            '지역': df['지역명'][:10],  # 상위 10개
            '유사도': df['유사도 점수'][:10]
        })
        chart_data.to_excel(writer, sheet_name='차트 데이터', index=False)
    
    def _create_details_sheet(self, writer, reference_area, recommendations):
        """상세 정보 시트 생성"""
        details_data = []
        for rec in recommendations:
            details_data.append({
                '지역명': rec.recommended_location.name,
                '시도': rec.recommended_location.sido,
                '시군구': rec.recommended_location.sigungu,
                '동': rec.recommended_location.dong,
                '총 인구': rec.recommended_location.population_total,
                '20대 인구': rec.recommended_location.population_20s,
                '30대 인구': rec.recommended_location.population_30s,
                '40대 인구': rec.recommended_location.population_40s,
                '50대 인구': rec.recommended_location.population_50s,
                '유동인구': rec.recommended_location.floating_population,
                '업종 밀도': rec.recommended_location.business_density,
                '공실률': rec.recommended_location.vacancy_rate,
                '상업지역 비율': rec.recommended_location.commercial_area_ratio,
                '주거지역 비율': rec.recommended_location.residential_area_ratio,
                '주차 가능성': rec.recommended_location.parking_availability_score
            })
        
        details_df = pd.DataFrame(details_data)
        details_df.to_excel(writer, sheet_name='상세 정보', index=False)
    
    def _apply_excel_styles(self, file_path):
        """Excel 스타일 적용"""
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 행 스타일 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    async def generate_pdf_report(
        self,
        reference_area: ReferenceArea,
        recommendations: List[RecommendationResult],
        include_charts: bool = True,
        include_details: bool = True
    ) -> str:
        """PDF 보고서 생성"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"franchise_recommendation_{reference_area.name}_{timestamp}.pdf"
        file_path = os.path.join(self.reports_dir, file_name)
        
        # PDF 문서 생성
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 제목
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1  # 가운데 정렬
        )
        story.append(Paragraph("미라트 스튜디오 가맹 추천 지역 분석 보고서", title_style))
        story.append(Spacer(1, 20))
        
        # 기준 상권 정보
        story.append(Paragraph("기준 상권 정보", styles['Heading2']))
        reference_info = [
            ['항목', '내용'],
            ['상권명', reference_area.name],
            ['주소', reference_area.address],
            ['월 매출', f"{reference_area.monthly_sales:,}원" if reference_area.monthly_sales else 'N/A'],
            ['상권 유형', reference_area.area_type or 'N/A'],
            ['분석 일시', datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        reference_table = Table(reference_info)
        reference_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(reference_table)
        story.append(Spacer(1, 20))
        
        # 추천 결과 요약
        story.append(Paragraph("추천 결과 요약", styles['Heading2']))
        
        if recommendations:
            avg_similarity = sum(float(r.similarity_score) for r in recommendations) / len(recommendations)
            max_similarity = max(float(r.similarity_score) for r in recommendations)
            reviewed_count = sum(1 for r in recommendations if r.is_reviewed)
            
            summary_info = [
                ['항목', '값'],
                ['총 추천 지역 수', str(len(recommendations))],
                ['평균 유사도', f"{avg_similarity:.3f}"],
                ['최고 유사도', f"{max_similarity:.3f}"],
                ['검토 완료', f"{reviewed_count}/{len(recommendations)}개"]
            ]
            
            summary_table = Table(summary_info)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
        
        # 상위 추천 지역 (최대 10개)
        story.append(Spacer(1, 20))
        story.append(Paragraph("상위 추천 지역", styles['Heading2']))
        
        if recommendations:
            top_recommendations = recommendations[:10]
            rec_data = [['순위', '지역명', '주소', '유사도', '추천 사유']]
            
            for i, rec in enumerate(top_recommendations, 1):
                rec_data.append([
                    str(i),
                    rec.recommended_location.name,
                    rec.recommended_location.address[:30] + '...' if len(rec.recommended_location.address) > 30 else rec.recommended_location.address,
                    f"{float(rec.similarity_score):.3f}",
                    rec.recommendation_reason[:20] + '...' if len(rec.recommendation_reason) > 20 else rec.recommendation_reason
                ])
            
            rec_table = Table(rec_data)
            rec_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(rec_table)
        
        # PDF 생성
        doc.build(story)
        
        return file_path
    
    async def generate_chart_image(self, recommendations: List[RecommendationResult]) -> str:
        """차트 이미지 생성"""
        if not recommendations:
            return None
        
        # 상위 10개 지역의 유사도 차트
        top_10 = recommendations[:10]
        locations = [r.recommended_location.name for r in top_10]
        scores = [float(r.similarity_score) for r in top_10]
        
        plt.figure(figsize=(12, 6))
        plt.bar(locations, scores, color='skyblue')
        plt.title('상위 추천 지역 유사도 점수', fontsize=14, fontweight='bold')
        plt.xlabel('지역명', fontsize=12)
        plt.ylabel('유사도 점수', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        
        # 이미지 저장
        chart_path = os.path.join(self.reports_dir, f"chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path